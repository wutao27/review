#!/usr/bin/env python

from openpyxl.reader.excel import load_workbook
import time
import pickle


def main():
    wb = load_workbook(filename = r'Reviewer report.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    row_range = ws.get_highest_row()
    column_range = ws.get_highest_column()

    origi_list = []

    for rx in range(row_range-1):
        print rx
        date_inv = ws.cell(row = rx+1,column = 0).value
        if date_inv!= None:
            temp_dic = {}
            for cx in range(column_range):
                item_name = ws.cell(row = 0,column = cx).value
                item_value = ws.cell(row = rx+1,column = cx).value
                temp_dic[item_name] = item_value
            origi_list.append(temp_dic)

    with open('origi_list.pickle','wb') as fid:
        pickle.dump(origi_list,fid)


if __name__ == '__main__':
    main()